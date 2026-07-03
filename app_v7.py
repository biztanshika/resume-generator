"""
Phase 7 -- Streamlit Resume Generator App (NEW FORMAT) + Role-Fit Analyzer.


Identical UI to app_v6 (upload -> OpenAI extract -> DOCX -> role-fit analysis).
The ONLY differences from app_v6 are:
  1. DOCX generation is routed through resume_generator_v2.create_resume,
     which renders the approved new-format design (blue header bar, navy
     headings, Professional Summary section, prose-style summary bullets,
     Responsibilities heading, horizontal rules, etc.).
  2. SYSTEM_PROMPT is tuned for the new format: prose-style summary bullets
     and explicit Professional Summary + Certifications extraction guidance.
The JSON schema returned by the model is unchanged from app_v6, so the v1
app (app_v6) and the v1 engine (resume_generator.py) keep working in parallel
and are not touched.
"""


import csv
import io
import json
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path


import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI


# NEW-FORMAT engine -- resume_generator_v2 is independent of the v1 module.
from resume_generator_v2 import create_resume


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
load_dotenv(Path(__file__).resolve().parent / ".env")


MODEL = "gpt-4.1"
MAX_PARALLEL_WORKERS = 10




def _resolve_api_key() -> str:
    """Return the active OpenAI key: UI-entered key > .env fallback."""
    return (
        st.session_state.get("user_api_key", "").strip()
        or os.getenv("OPENAI_API_KEY", "")
    )


# ---------------------------------------------------------------------------
# WorkPace-inspired theme CSS
# ---------------------------------------------------------------------------
WORKPACE_CSS = """
<style>
    /* ----- global background ----- */
    .stApp {
        background-color: #EAF4FB;
    }


    /* ----- header area ----- */
    header[data-testid="stHeader"] {
        background-color: #EAF4FB;
    }


    /* ----- card container ----- */
    .wp-card {
        background: #ffffff;
        border-radius: 12px;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        padding: 2rem;
        margin-bottom: 1.5rem;
    }


    /* ----- brand header ----- */
    .wp-brand {
        font-size: 2.2rem;
        font-weight: 700;
        color: #0A3D62;
        margin-bottom: 0.25rem;
    }
    .wp-brand span {
        color: #0077B6;
    }


    .wp-subtitle {
        color: #555;
        font-size: 1.05rem;
        margin-bottom: 1.5rem;
    }


    /* ----- buttons ----- */
    .stButton > button {
        background-color: #0A3D62 !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.6rem 2rem !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        transition: background 0.2s;
    }
    .stButton > button:hover {
        background-color: #0077B6 !important;
    }


    /* ----- download button ----- */
    .stDownloadButton > button {
        background-color: #0077B6 !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.6rem 2rem !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
    }
    .stDownloadButton > button:hover {
        background-color: #0A3D62 !important;
    }


    /* ----- file uploader zone ----- */
    [data-testid="stFileUploader"] {
        border: 2px dashed #0077B6;
        border-radius: 12px;
        padding: 1rem;
        background-color: #ffffff;
    }
    [data-testid="stFileUploaderDropzone"] {
        background-color: #F0F7FC !important;
        border: none !important;
        border-radius: 8px !important;
    }
    [data-testid="stFileUploaderDropzone"] span,
    [data-testid="stFileUploaderDropzone"] small,
    [data-testid="stFileUploaderDropzone"] p,
    [data-testid="stFileUploaderDropzone"] div {
        color: #0A3D62 !important;
    }
    [data-testid="stFileUploaderDropzone"] button {
        color: #0A3D62 !important;
        border-color: #0077B6 !important;
        background-color: #ffffff !important;
    }
    [data-testid="stFileUploaderDropzone"] svg {
        fill: #0077B6 !important;
        stroke: #0077B6 !important;
    }


    /* ----- spinner ----- */
    .stSpinner > div {
        border-top-color: #0077B6 !important;
    }


    /* ----- success / error ----- */
    .stAlert {
        border-radius: 8px;
    }


    /* ----- widget labels (file uploader, text area, etc.) ----- */
    [data-testid="stWidgetLabel"] p {
        color: #0A3D62 !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
    }


    /* ----- bordered containers (upload card, processing card) ----- */
    [data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #ffffff !important;
        border-radius: 12px !important;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08) !important;
        border: none !important;
        padding: 0.5rem !important;
    }


    /* ----- text area white background ----- */
    [data-testid="stTextArea"] textarea {
        background-color: #ffffff !important;
    }


    /* ----- Phase 5: analysis canvas (right panel) ----- */
    .analysis-canvas {
        background: #FFFFFF;
        border-radius: 12px;
        border: 1px solid #E2E8F0;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        padding: 1.5rem 1.75rem;
    }


    /* Score row */
    .canvas-score-row {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding-bottom: 0.9rem;
        margin-bottom: 1rem;
        border-bottom: 1px solid #E2E8F0;
    }
    .canvas-score-label {
        font-size: 1.25rem;
        font-weight: 700;
        color: #1E293B;
    }
    .score-badge {
        display: inline-block;
        padding: 0.2rem 0.75rem;
        border-radius: 999px;
        font-size: 1rem;
        font-weight: 700;
    }
    .score-badge-high { background: #ECFDF5; color: #065F46; }
    .score-badge-mid  { background: #FEF9EC; color: #D97706; }
    .score-badge-low  { background: #FEF2F2; color: #DC2626; }


    /* Summary strip */
    .fit-summary {
        background: #F1F5F9;
        border-left: 3px solid #94A3B8;
        border-radius: 4px;
        padding: 0.6rem 0.9rem;
        color: #475569;
        font-size: 0.9rem;
        font-style: italic;
        margin-bottom: 1.1rem;
    }


    /* Section blocks */
    .fit-block {
        border-radius: 8px;
        padding: 0.9rem 1.15rem;
        margin-bottom: 0.9rem;
        border-left: 4px solid;
    }
    .fit-block-strengths   { background: #ECFDF5; border-color: #10B981; }
    .fit-block-weaknesses  { background: #FEF2F2; border-color: #EF4444; }
    .fit-block-suggestions { background: #EFF6FF; border-color: #3B82F6; }


    .fit-block-title { font-size: 0.95rem; font-weight: 700; margin-bottom: 0.5rem; }
    .fit-title-strength   { color: #065F46; }
    .fit-title-weakness   { color: #7F1D1D; }
    .fit-title-suggestion { color: #1E3A8A; }


    .fit-block-item { font-size: 0.88rem; line-height: 1.65; margin-bottom: 0.25rem; }
    .fit-item-strength   { color: #065F46; }
    .fit-item-weakness   { color: #7F1D1D; }
    .fit-item-suggestion { color: #1E3A8A; }
</style>
"""


# ---------------------------------------------------------------------------
# AI system prompt  (derived from the Custom GPT prompt, outputs JSON)
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """\
You are a Resume Content Extractor.


You will receive raw resume text wrapped in <resume>...</resume> delimiters \
(from a .docx, .pdf, or .txt upload). Your job is to parse and reorganise that \
content into a structured JSON object that matches the schema below. Return \
ONLY valid JSON -- no markdown, no explanation, no extra keys.


### Your Single Most Important Rule


Your job is FAITHFUL PRESERVATION of what the candidate wrote -- not helpful \
improvement. Do NOT shorten, summarise, consolidate, merge, paraphrase down, \
or make the resume "tighter." Equally: do NOT pad, invent, fabricate, or add \
any detail the candidate did not write. The output should MIRROR the source's \
verbosity: rich and detailed when the source is rich and detailed; brief when \
the source is brief. Your task is reorganisation, not editing.


### Core Rules (apply to every section)


- Use ONLY information explicitly present in the resume between the <resume> \
  tags. Never invent, infer, or add any detail not stated by the candidate.
- Preserve content depth in proportion to the source. If a role has 8 bullets \
  in the resume, output 8 bullets. If a bullet is 30 words, keep it near 30 \
  words. If the resume is sparse, the output is sparse -- do NOT pad it.
- Light rewording for active voice and strong verbs is allowed; structural \
  shortening, merging two bullets into one, or compressing a 30-word bullet \
  to 12 words is NOT allowed.
- Specific numbers, percentages, dollar amounts, dates, tool names, and \
  acronyms must be preserved EXACTLY as written.


### Content Rules


1. **name** -- The candidate's full name.


2. **summary** -- A list of 6-8 prose-style bullet strings forming a \
"Professional Summary" paragraph set. Each bullet should be a flowing, \
complete sentence. For each bullet:
   - Write 1-2 complete sentences (typically 25-60 words) in active voice.
   - Do NOT prefix with a bold label or colon. The first word should be a \
     descriptive adjective, role noun, or verb (e.g. "Results-driven Data \
     Engineer with...", "Proven track record of...", "Domain experience spans...", \
     "Demonstrated leadership in...").
   - Cover the candidate's experience breadth, tooling depth, domain expertise, \
     leadership/collaboration, and signature achievements -- one theme per bullet.
   - Derive ALL content from what is actually written in the resume; do not pad \
     with generic phrases.
   - If the source resume has a dedicated "Professional Summary" / "Profile" / \
     "About" / "Career Summary" section, use it as the primary source and \
     expand into 6-8 prose bullets. If no such section exists, synthesise the \
     bullets from the candidate's roles, skills, and projects.
   - If the resume genuinely lacks enough material for 6 bullets, produce \
     however many you can fully back with source content. Never invent themes.


3. **experience** -- A list of objects, one per role:
   - "role": Job title ONLY -- strip company name, location, and dates.
   - "tech": Comma-separated list of technologies/tools used in that role, \
     taken directly from the resume text for that role. Do not add any \
     technology not specifically mentioned in that role.
   - "bullets": Output EVERY duty and accomplishment bullet from the original \
     resume for this role, in the same order. Do not drop, merge, or combine \
     bullets. Reword each bullet using active voice and strong verbs, but \
     keep its length and specificity faithful to the source. A 30-word source \
     bullet stays roughly 30 words; a 10-word source bullet stays roughly 10 \
     words. NEVER pad short bullets to make them look longer.


     BAD (compression -- forbidden):
       Source: "Designed and implemented a multi-tenant Azure Data Factory \
       pipeline orchestrating 40+ daily incremental loads from Salesforce, \
       NetSuite, and on-prem SQL Server into a Synapse star schema, cutting \
       reporting latency from 6 hours to 22 minutes."
       Output: "Built Azure ETL pipelines that improved reporting speed."


     GOOD (faithful preservation -- required):
       Source: "Designed and implemented a multi-tenant Azure Data Factory \
       pipeline orchestrating 40+ daily incremental loads from Salesforce, \
       NetSuite, and on-prem SQL Server into a Synapse star schema, cutting \
       reporting latency from 6 hours to 22 minutes."
       Output: "Designed a multi-tenant Azure Data Factory pipeline \
       orchestrating 40+ daily incremental loads from Salesforce, NetSuite, \
       and on-prem SQL Server into a Synapse star schema, reducing reporting \
       latency from 6 hours to 22 minutes."


4. **projects** -- A list of objects, one per project:
   - "title": Project name ONLY -- strip technology tags, locations, dates, and \
     any company or client names.
   - "bullets": Preserve ALL detail bullets from the original resume for this \
     project, faithfully (same rules as experience bullets above). Remove any \
     company or client names from bullet text -- replace them with a neutral \
     descriptor (e.g. "the client", "a financial services firm") only if \
     removal would make the sentence grammatically incomplete; otherwise \
     simply omit the name. Each bullet should capture technologies used, \
     outcomes achieved, and responsibilities held, at the original level of \
     detail.


5. **education** -- A list of objects, one per degree. If the resume lists \
multiple degrees (e.g. a Bachelor's AND a Master's), you MUST include a \
separate object for EACH degree. Never combine or drop degrees. Each object has:
   - "degree": Degree name.
   - "college": Institution name ONLY -- do not include graduation year or dates.


6. **skills** -- A list of strings, each in the format \
"Category: item1, item2, item3". Group by meaningful categories based on \
resume content (e.g. Programming, Web Technologies, Databases, Tools, \
Methodologies, Cloud, BI Tools, Collaboration & Leadership, Domain Expertise, \
Problem-Solving, etc.). Include all skills mentioned in the resume.


7. **certifications** -- A list of certification strings. Each entry should be \
written in "Issuer: Certification Name" format when an issuer is identifiable \
(e.g. "Microsoft: Azure Solutions Architect Expert (AZ-305)", "AWS Certified: \
Solutions Architect - Associate"). If the resume has a dedicated \
"Certifications" / "Licenses & Certifications" / "Professional Credentials" \
section, extract EVERY entry verbatim (preserve issuer, name, and exam code). \
If the resume mentions certifications inline within other sections, surface \
them here too. Return an empty list [] only if the resume mentions no \
certifications at all.


### Worked Example (Experience entry)


If the source <resume> contained this role:


  Senior Data Engineer, Acme Corp -- 2019-2023
  Tools: Python, Airflow, Snowflake, dbt, Looker
   * Architected and rolled out a Snowflake + dbt warehouse serving 14 \
     analytics teams; modelled 80+ business entities and cut nightly batch \
     time from 4.5 hours to 38 minutes.
   * Led migration of 120 legacy Informatica jobs to Airflow DAGs, \
     consolidating three orchestration tools and saving ~$140K in annual \
     licensing.
   * Mentored four junior engineers, instituted weekly code reviews, and \
     authored the team's testing playbook (pytest + Great Expectations).
   * Partnered with finance to deliver a self-serve revenue dashboard \
     adopted by 60+ users in the first quarter.


The correct extracted entry would be:


  {
    "role": "Senior Data Engineer",
    "tech": "Python, Airflow, Snowflake, dbt, Looker",
    "bullets": [
      "Architected and rolled out a Snowflake + dbt warehouse serving 14 \
analytics teams, modelling 80+ business entities and cutting nightly batch \
time from 4.5 hours to 38 minutes.",
      "Led migration of 120 legacy Informatica jobs to Airflow DAGs, \
consolidating three orchestration tools and saving roughly $140K in annual \
licensing.",
      "Mentored four junior engineers, instituted weekly code reviews, and \
authored the team's testing playbook using pytest and Great Expectations.",
      "Partnered with finance to deliver a self-serve revenue dashboard \
adopted by 60+ users in the first quarter."
    ]
  }


Notice: company name and dates stripped; tools preserved exactly; all four \
bullets kept; numbers and tool names preserved verbatim; lengths roughly \
match the source. Nothing added, nothing dropped.


### JSON Schema


{
  "name": "string",
  "summary": ["prose-style sentence 1", "prose-style sentence 2", "..."],
  "experience": [
    {
      "role": "string",
      "tech": "string",
      "bullets": ["string"]
    }
  ],
  "projects": [
    {
      "title": "string",
      "bullets": ["string"]
    }
  ],
  "education": [
    {
      "degree": "string",
      "college": "string"
    },
    {
      "degree": "string (include ALL degrees)",
      "college": "string"
    }
  ],
  "skills": ["Category: items", "..."],
  "certifications": ["Issuer: Certification Name", "..."]
}


### Final Reminder (read this before you start)


Faithful preservation, not helpful improvement. If the source is detailed, \
your output is detailed. If the source is sparse, your output is sparse. \
Never drop bullets. Never merge bullets. Never compress a long bullet into \
a short one. Never invent or pad a short bullet into a long one. Strip \
company/client names from project titles and project bullets. Strip \
graduation years from education. Return ONLY the JSON object -- no markdown \
fences, no commentary.\
"""


# ---------------------------------------------------------------------------
# Phase 3: Role-fit rating system prompt
# ---------------------------------------------------------------------------
RATING_SYSTEM_PROMPT = """\
You are a Resume-to-Job Fit Analyst.


You will receive two pieces of text:
1. **RESUME** -- the full text of a candidate's resume.
2. **JOB CONTEXT** -- a free-form description provided by the user. It may \
include a job title, project description, required technologies, tools, \
responsibilities, team structure, domain, or any combination of these.


Your task is to evaluate how well the resume matches the job context and return \
a structured JSON assessment. Return ONLY valid JSON -- no markdown, no \
explanation, no extra keys.


### Evaluation Criteria


Score the resume on a 1-10 scale considering ALL of the following:
- Skills match: Do the candidate's skills, tools, and technologies align with \
  what the job context requires?
- Experience relevance: Does the work history show relevant responsibilities, \
  domains, or project types?
- Project alignment: Do projects use similar technologies or solve similar \
  problems?
- Keyword coverage: Are the specific tools, frameworks, methodologies, and \
  domain terms from the job context present in the resume?
- Seniority fit: Does the experience level match what the role requires?


### Output Rules


1. **strengths** -- List 3-6 specific things that make the resume a good fit. \
Reference actual resume content (e.g. "3 years of Python experience aligns \
with the required Python proficiency"). Each item is a plain string.


2. **weaknesses** -- List every significant gap or shortcoming. This includes:
   - Missing skills, tools, or technologies mentioned in the job context but \
     absent from the resume.
   - Experience gaps -- areas where the candidate's experience is thin, \
     irrelevant, or too junior relative to the role.
   - Domain mismatches or missing certifications.
   Each item is a plain string that clearly states what is missing or weak.


3. **suggestions** -- List 3-6 concrete, actionable steps the candidate should \
take to better align their resume with this specific job context. Focus on \
content changes: which sections to strengthen, what keywords to add, which \
experience to highlight or reframe. Do NOT suggest formatting or layout changes. \
Each item is a plain string.


### JSON Schema


{
  "overall_score": <integer 1-10>,
  "score_out_of": 10,
  "summary": "<1-2 sentence overall verdict>",
  "strengths": ["<string>", "..."],
  "weaknesses": ["<string>", "..."],
  "suggestions": ["<string>", "..."]
}


Return ONLY the JSON object. No markdown fences, no commentary.\
"""


# ---------------------------------------------------------------------------
# Text extraction helpers  (inline, no separate module)
# ---------------------------------------------------------------------------


def extract_text_docx(file_bytes: bytes) -> str:
    """Extract text from a .docx file using python-docx."""
    from docx import Document as DocxDocument


    with io.BytesIO(file_bytes) as buf:
        doc = DocxDocument(buf)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_text_pdf(file_bytes: bytes) -> str:
    """Extract text from a PDF using pdfplumber."""
    import pdfplumber


    text_parts = []
    with io.BytesIO(file_bytes) as buf:
        with pdfplumber.open(buf) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
    return "\n".join(text_parts)


def extract_text_txt(file_bytes: bytes) -> str:
    """Decode a plain-text file."""
    return file_bytes.decode("utf-8", errors="replace")


EXTRACTORS = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": extract_text_docx,
    "application/pdf": extract_text_pdf,
    "text/plain": extract_text_txt,
}


def extract_text(uploaded_file) -> str:
    """Route to the correct extractor based on MIME type."""
    mime = uploaded_file.type
    file_bytes = uploaded_file.getvalue()


    # Fallback by extension if MIME is generic
    if mime not in EXTRACTORS:
        name_lower = uploaded_file.name.lower()
        if name_lower.endswith(".docx"):
            return extract_text_docx(file_bytes)
        if name_lower.endswith(".pdf"):
            return extract_text_pdf(file_bytes)
        if name_lower.endswith(".txt"):
            return extract_text_txt(file_bytes)
        raise ValueError(f"Unsupported file type: {mime} ({uploaded_file.name})")


    return EXTRACTORS[mime](file_bytes)




def _sanitize_text(text: str) -> str:
    """Replace common non-ASCII characters so ASCII-only encoding paths don't fail."""
    return (
        text
        .replace("\u2014", "--")   # em dash
        .replace("\u2013", "-")    # en dash
        .replace("\u2018", "'")    # left single quote
        .replace("\u2019", "'")    # right single quote
        .replace("\u201c", '"')   # left double quote
        .replace("\u201d", '"')   # right double quote
        .replace("\u2026", "...")  # ellipsis
        .replace("\u00a0", " ")   # non-breaking space
        .replace("\u2022", "-")   # bullet
        .replace("\u2010", "-")   # hyphen
        .replace("\u2011", "-")   # non-breaking hyphen
    )


# ---------------------------------------------------------------------------
# OpenAI calls
# ---------------------------------------------------------------------------
REQUIRED_KEYS = {"name", "summary", "experience", "projects", "education", "skills"}


def call_openai(resume_text: str, api_key: str = "") -> dict:
    """Send resume text to the model and return the parsed JSON dict."""
    key = api_key or _resolve_api_key()
    if not key:
        raise RuntimeError("Please enter your OpenAI API key above.")


    client = OpenAI(api_key=key)


    # Wrap resume in explicit delimiters so the model never confuses content
    # with instructions, and the boundary is unambiguous.
    user_message = f"<resume>\n{resume_text}\n</resume>"


    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_message},
        ],
        response_format={"type": "json_object"},
        temperature=0.3,
    )


    raw = response.choices[0].message.content
    data = json.loads(raw)


    # Validate required keys
    missing = REQUIRED_KEYS - set(data.keys())
    if missing:
        raise ValueError(f"AI response missing required keys: {missing}")


    # Ensure certifications key exists (may be absent)
    data.setdefault("certifications", [])


    return data


RATING_REQUIRED_KEYS = {
    "overall_score", "score_out_of", "summary",
    "strengths", "weaknesses", "suggestions",
}


def call_openai_rating(resume_text: str, job_context: str, api_key: str = "") -> dict:
    """Rate the resume against a free-form job context description."""
    key = api_key or _resolve_api_key()
    if not key:
        raise RuntimeError("Please enter your OpenAI API key above.")


    user_message = (
        "=== RESUME ===\n"
        f"{resume_text}\n\n"
        "=== JOB CONTEXT ===\n"
        f"{job_context}"
    )


    client = OpenAI(api_key=key)


    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": RATING_SYSTEM_PROMPT},
            {"role": "user", "content": user_message},
        ],
        response_format={"type": "json_object"},
        temperature=0.4,
    )


    raw = response.choices[0].message.content
    data = json.loads(raw)


    missing = RATING_REQUIRED_KEYS - set(data.keys())
    if missing:
        raise ValueError(f"Rating response missing required keys: {missing}")


    return data


# ---------------------------------------------------------------------------
# DOCX generation wrapper  (writes to BytesIO buffer)
# ---------------------------------------------------------------------------


def generate_resume_bytes(data: dict) -> bytes:
    """Call resume_generator_v2.create_resume() and return the .docx bytes."""
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp_path = tmp.name


    try:
        create_resume(data, output_path=tmp_path)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ---------------------------------------------------------------------------
# Phase 3: Rating display helpers
# ---------------------------------------------------------------------------


def _score_badge_class(score: int) -> str:
    """Return badge CSS class based on score."""
    if score >= 8:
        return "score-badge score-badge-high"
    if score >= 5:
        return "score-badge score-badge-mid"
    return "score-badge score-badge-low"


def display_rating_canvas(rating: dict):
    """Render the 3-section role-fit analysis as left-border accent blocks."""
    score   = rating.get("overall_score", 0)
    out_of  = rating.get("score_out_of", 10)
    badge   = _score_badge_class(score)
    summary = rating.get("summary", "")


    parts = ['<div class="analysis-canvas">']


    # Score row with badge pill
    parts.append(
        f'<div class="canvas-score-row">'
        f'<span class="canvas-score-label">Score</span>'
        f'<span class="{badge}">{score}/{out_of}</span>'
        f'</div>'
    )


    # Summary strip
    if summary:
        parts.append(f'<div class="fit-summary">{summary}</div>')


    # Strengths block
    strengths = rating.get("strengths", [])
    if strengths:
        items = "".join(
            f'<div class="fit-block-item fit-item-strength">&bull; {s}</div>'
            for s in strengths
        )
        parts.append(
            f'<div class="fit-block fit-block-strengths">'
            f'<div class="fit-block-title fit-title-strength">Strength</div>'
            f'{items}</div>'
        )


    # Weaknesses block
    weaknesses = rating.get("weaknesses", [])
    if weaknesses:
        items = "".join(
            f'<div class="fit-block-item fit-item-weakness">&bull; {w}</div>'
            for w in weaknesses
        )
        parts.append(
            f'<div class="fit-block fit-block-weaknesses">'
            f'<div class="fit-block-title fit-title-weakness">Weaknesses</div>'
            f'{items}</div>'
        )


    # Suggestions block
    suggestions = rating.get("suggestions", [])
    if suggestions:
        items = "".join(
            f'<div class="fit-block-item fit-item-suggestion">&bull; {s}</div>'
            for s in suggestions
        )
        parts.append(
            f'<div class="fit-block fit-block-suggestions">'
            f'<div class="fit-block-title fit-title-suggestion">Suggestions</div>'
            f'{items}</div>'
        )


    parts.append('</div>')
    st.markdown("".join(parts), unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Bulk analysis helpers
# ---------------------------------------------------------------------------


def _extract_job_context(uploaded_jd) -> str:
    """Extract plain text from an uploaded job description file."""
    if uploaded_jd is None:
        return ""
    mime = uploaded_jd.type
    file_bytes = uploaded_jd.getvalue()
    name_lower = uploaded_jd.name.lower()
    if mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or name_lower.endswith(".docx"):
        return extract_text_docx(file_bytes)
    if mime == "application/pdf" or name_lower.endswith(".pdf"):
        return extract_text_pdf(file_bytes)
    return file_bytes.decode("utf-8", errors="replace")




def _analyse_one(filename, file_bytes, file_mime, job_context, api_key):
    """Analyse a single resume for job fit. Runs in a worker thread."""
    class _FakeFile:
        def __init__(self, name, data, mime):
            self.name = name
            self.type = mime
            self._data = data
        def getvalue(self):
            return self._data


    try:
        text = _sanitize_text(extract_text(_FakeFile(filename, file_bytes, file_mime)))
        if not text.strip():
            return {"filename": filename, "error": "File appears to be empty", "rating": None}
        rating = call_openai_rating(text, job_context, api_key=api_key)
        return {"filename": filename, "error": None, "rating": rating}
    except Exception as exc:
        return {"filename": filename, "error": str(exc), "rating": None}




def _build_csv_bytes(results):
    """Build a UTF-8 CSV summary from bulk analysis results."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Rank", "File", "Score", "Verdict",
        "Top Strength", "Top Weakness", "Top Suggestion",
    ])
    for i, r in enumerate(results, 1):
        rt = r["rating"] or {}
        writer.writerow([
            i,
            r["filename"],
            rt.get("overall_score", "Error"),
            rt.get("summary", r.get("error", "")),
            next(iter(rt.get("strengths", [])), ""),
            next(iter(rt.get("weaknesses", [])), ""),
            next(iter(rt.get("suggestions", [])), ""),
        ])
    return buf.getvalue().encode("utf-8")




def _build_excel_bytes(results):
    """Build a styled Excel workbook. Returns None if openpyxl is unavailable."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fit Analysis"


    headers = ["Rank", "File", "Score", "Verdict", "Strengths", "Weaknesses", "Suggestions"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A3D62")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


    for i, r in enumerate(results, 1):
        rt = r["rating"] or {}
        ws.append([
            i,
            r["filename"],
            rt.get("overall_score", "Error"),
            rt.get("summary", r.get("error", "")),
            "\n".join(f"- {s}" for s in rt.get("strengths", [])),
            "\n".join(f"- {s}" for s in rt.get("weaknesses", [])),
            "\n".join(f"- {s}" for s in rt.get("suggestions", [])),
        ])
        for cell in ws[i + 1]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


    for col_idx, width in enumerate([6, 32, 8, 52, 52, 52, 52], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()




def _render_bulk_results(results):
    """Render bulk fit analysis as a ranked expandable list with downloads."""
    n_ok = sum(1 for r in results if r["rating"] is not None)
    n_err = len(results) - n_ok
    err_text = f" -- {n_err} failed" if n_err else ""
    st.markdown(
        '<div class="wp-card">'
        '<div class="wp-brand" style="font-size:1.4rem;">Bulk Fit Analysis</div>'
        '<div class="wp-subtitle">'
        f"{n_ok} of {len(results)} resumes analysed{err_text}, ranked by fit score."
        "</div>"
        "</div>",
        unsafe_allow_html=True,
    )


    dl1, dl2, _ = st.columns([1, 1, 5])
    with dl1:
        st.download_button(
            "Download CSV",
            data=_build_csv_bytes(results),
            file_name="fit_analysis.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with dl2:
        excel_bytes = _build_excel_bytes(results)
        if excel_bytes is not None:
            st.download_button(
                "Download Excel",
                data=excel_bytes,
                file_name="fit_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


    st.write("")
    for i, result in enumerate(results, 1):
        rating = result["rating"]
        if rating:
            score = rating.get("overall_score", 0)
            verdict = rating.get("summary", "")
            label = f"#{i}  {result['filename']}  --  {score}/10  |  {verdict}"
        else:
            label = f"#{i}  {result['filename']}  --  Failed"


        with st.expander(label, expanded=(i == 1)):
            if rating:
                display_rating_canvas(rating)
            else:
                st.error(f"Could not analyse this resume: {result['error']}")




# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------


def main():
    st.set_page_config(
        page_title="Resume Generator",
        page_icon="briefcase",
        layout="wide",
    )


    # Inject WorkPace CSS
    st.markdown(WORKPACE_CSS, unsafe_allow_html=True)


    # Decide layout: centered single column by default,
    # left + right panel when analysis data exists.
    has_rating = "rating_data" in st.session_state


    if has_rating:
        left_col, right_col = st.columns([4, 5])
    else:
        # Simulate a centered column inside wide layout
        _pad_l, left_col, _pad_r = st.columns([1, 3, 1])
        right_col = None


    # Track whether we need to rerun after column blocks close
    needs_rerun = False


    # ---- Left column: brand + upload + download ----
    with left_col:
        # ---- Brand header ----
        st.markdown(
            '<div class="wp-card">'
            '  <div class="wp-brand">Resume<span>Gen</span></div>'
            '  <div class="wp-subtitle">'
            "    Upload your resume and get a professionally formatted Word document "
            "    powered by AI."
            "  </div>"
            "</div>",
            unsafe_allow_html=True,
        )


        # ---- API key input ----
        api_key_input = st.text_input(
            "OpenAI / GPT API Key",
            type="password",
            placeholder="sk-...",
            help=(
                "Enter your OpenAI API key. It is used only for this session "
                "to process your resume via the GPT API and is not stored."
            ),
        ).strip()
        if api_key_input:
            st.session_state["user_api_key"] = api_key_input


        # ---- Upload section ----
        with st.container(border=True):
            uploaded_files = st.file_uploader(
                "Drop resume(s) here",
                type=["docx", "pdf", "txt"],
                accept_multiple_files=True,
                help="Single file: all three actions available. Multiple files: Analyse Fit only.",
            ) or []
            has_any_file = len(uploaded_files) > 0
            is_multiple = len(uploaded_files) > 1
            if is_multiple:
                st.caption(
                    f"{len(uploaded_files)} files selected -- "
                    "Build Resume and Build & Analyse support single files only."
                )


            jd_tab_type, jd_tab_upload = st.tabs(["Type / Paste", "Upload File"])
            with jd_tab_type:
                job_context_typed = st.text_area(
                    "Job Description / Role Context (optional)",
                    placeholder=(
                        "Describe the role, project, required technologies, "
                        "responsibilities, and any specific skills needed. "
                        "Leave blank to skip the role-fit analysis."
                    ),
                    height=120,
                    label_visibility="collapsed",
                )
            with jd_tab_upload:
                jd_file = st.file_uploader(
                    "Upload job description",
                    type=["pdf", "docx", "txt"],
                    help="PDF, DOCX, or TXT containing the job description.",
                    key="jd_uploader",
                    label_visibility="collapsed",
                )
                if jd_file:
                    st.caption(f"Loaded: {jd_file.name}")
            job_context = _extract_job_context(jd_file) if jd_file else job_context_typed


            # Three independent action buttons
            btn1, btn2, btn3 = st.columns(3)
            with btn1:
                gen_clicked = st.button(
                    "Build Resume",
                    disabled=not has_any_file or is_multiple,
                    use_container_width=True,
                )
            with btn2:
                fit_clicked = st.button(
                    "Analyse Fit",
                    disabled=not has_any_file,
                    use_container_width=True,
                )
            with btn3:
                both_clicked = st.button(
                    "Build & Analyse",
                    disabled=not has_any_file or is_multiple,
                    use_container_width=True,
                )


        # ---- Processing (runs on button click, stores results in session) ----
        if (gen_clicked or fit_clicked or both_clicked) and has_any_file:
            active_key = _resolve_api_key()
            if not active_key:
                st.error("Please enter your OpenAI API key above before proceeding.")
            else:
                # --- Bulk fit analysis (multiple files, Analyse Fit only) ---
                if fit_clicked and is_multiple:
                    if not job_context.strip():
                        st.warning(
                            "Please enter a job description or role context to run "
                            "the fit analysis."
                        )
                    else:
                        st.session_state.pop("rating_data", None)
                        file_data = [
                            (f.name, f.getvalue(), f.type) for f in uploaded_files
                        ]
                        progress_bar = st.progress(
                            0, text=f"Analysing... 0/{len(file_data)} resumes"
                        )
                        raw_results = [None] * len(file_data)
                        completed = 0
                        with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
                            future_to_idx = {
                                executor.submit(
                                    _analyse_one,
                                    fname, fbytes, ftype,
                                    job_context.strip(), active_key,
                                ): idx
                                for idx, (fname, fbytes, ftype) in enumerate(file_data)
                            }
                            for future in as_completed(future_to_idx):
                                idx = future_to_idx[future]
                                raw_results[idx] = future.result()
                                completed += 1
                                progress_bar.progress(
                                    completed / len(file_data),
                                    text=f"Analysing... {completed}/{len(file_data)} resumes",
                                )
                        progress_bar.empty()
                        st.session_state["bulk_results"] = sorted(
                            raw_results,
                            key=lambda r: (
                                r["rating"]["overall_score"] if r["rating"] else -1
                            ),
                            reverse=True,
                        )


                # --- Single-file paths ---
                else:
                    single_file = uploaded_files[0]
                    with st.spinner("Extracting text from your resume..."):
                        try:
                            resume_text = _sanitize_text(extract_text(single_file))
                        except Exception as exc:
                            st.error(f"Could not read the uploaded file: {exc}")
                            return


                    if not resume_text.strip():
                        st.error("The uploaded file appears to be empty.")
                        return


                    if gen_clicked or both_clicked:
                        with st.spinner("AI is analysing your resume..."):
                            try:
                                resume_data = call_openai(resume_text, api_key=active_key)
                            except Exception as exc:
                                st.error(f"AI processing failed: {exc}")
                                return


                        with st.spinner("Building your formatted resume..."):
                            try:
                                docx_bytes = generate_resume_bytes(resume_data)
                            except Exception as exc:
                                st.error(f"DOCX generation failed: {exc}")
                                return


                        safe_name = resume_data.get("name", "resume").replace(" ", "_")
                        st.session_state["docx_bytes"] = docx_bytes
                        st.session_state["docx_filename"] = f"{safe_name}_Resume.docx"


                    if fit_clicked or both_clicked:
                        if not job_context.strip():
                            st.warning(
                                "Please enter a job description or role context to run "
                                "the fit analysis."
                            )
                        else:
                            st.session_state.pop("bulk_results", None)
                            with st.spinner("Analysing role fit..."):
                                try:
                                    rating_data = call_openai_rating(
                                        resume_text, job_context.strip(),
                                        api_key=active_key,
                                    )
                                    st.session_state["rating_data"] = rating_data
                                    needs_rerun = True
                                except Exception as exc:
                                    st.error(f"Role-fit analysis failed: {exc}")
                                    st.session_state.pop("rating_data", None)


        # ---- Download button (persists via session state) ----
        if "docx_bytes" in st.session_state:
            with st.container(border=True):
                st.success("Your resume is ready!")
                st.download_button(
                    label="Download Resume",
                    data=st.session_state["docx_bytes"],
                    file_name=st.session_state["docx_filename"],
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                )


    # ---- Right column: single-resume analysis canvas ----
    if has_rating and right_col is not None:
        with right_col:
            display_rating_canvas(st.session_state["rating_data"])


    # ---- Bulk analysis results (full width, below upload area) ----
    if "bulk_results" in st.session_state:
        _render_bulk_results(st.session_state["bulk_results"])


    # Rerun AFTER all column blocks are closed to avoid ASGI crash
    # (RerunException inside a with-column context breaks some Streamlit versions)
    if needs_rerun:
        st.rerun()


if __name__ == "__main__":
    main()



